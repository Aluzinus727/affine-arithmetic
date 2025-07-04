import random
import numpy as np
import pandas as pd
import sympy as sp

from memory import Affine
from pyibex import Interval
from solvers.variant import LinearExpression, LinearEnvelope

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook

import numpy as np
from collections import defaultdict

class CircularQueue:
    def __init__(self):
        self.elements = []
        self.index = 0
        self.full = False

    def add(self, item):
        """
        Agrega un nuevo elemento al final de la cola.
        """
        self.elements.append(item)

    def get_next(self):
        """
        Devuelve el siguiente elemento de forma circular.
        Si no hay elementos, devuelve None.
        """
        if not self.elements or not self.full:
            return None

        item = self.elements[self.index]
        self.index = (self.index + 1) % len(self.elements)
        return item

    def get_all(self):
        """
        Devuelve una copia de todos los elementos.
        """
        return self.elements

    def reset(self):
        """
        Vacía la cola y reinicia el índice.
        """
        self.elements = []
        self.index = 0

    def tensors(self):
        return [item['tensor'] for item in self.elements]

    def __str__(self):
        return str(self.elements)


circular_queue = CircularQueue()

def process_data(new_results):
    raw_data = []  
    error_by_test_case = []
    error_summary = {}
    global_method_errors = {
        'AF1': [],
        'AF2': [],
        'Ibex': [],
        'RevisedAF': []
    }
    for fname, intervals_obj in new_results.items():
        for interval, implementations in intervals_obj.items():
            for implementation, info in implementations.items():
                data = {
                    'fname': fname,
                    'function': info['Function'],
                    'interval': info['Interval'],
                    'implementation': info['Implementation'],
                    'LB': info['Min'],
                    'UB': info['Max']
                }
                raw_data.append(data)

    for fname, intervals_obj in new_results.items():
        method_errors = {
            'AF1': [],
            'AF2': [],
            'RevisedAF': [],
            'Ibex': []
        }
        for interval, implementations in intervals_obj.items():
            ranges = []
            for implementation, info in implementations.items():
                ranges.append({
                    'implementation': implementation,
                    'range': info['Max'] - info['Min']
                })
            
            best_range = min(r['range'] for r in ranges)
            for r in ranges:
                err = 0.0 if r["range"] == best_range else ((r["range"] - best_range) / best_range) * 100
                method_errors[r["implementation"]].append(err)
            
        avg_errors = {method: round(sum(errors) / len(errors), 2) for method, errors in method_errors.items()}
        error_summary[fname] = avg_errors

    for fname, intervals in new_results.items():
        for interval_data in intervals.values():
            ranges = [
                {
                    "implementation": impl,
                    "range": data["Max"] - data["Min"]
                }
                for impl, data in interval_data.items()
            ]
            best_range = min(r["range"] for r in ranges)

            for r in ranges:
                err = 0.0 if r["range"] == best_range else ((r["range"] - best_range) / best_range) * 100
                global_method_errors[r["implementation"]].append(err)

    # Calcular promedio y desviación estándar con numpy
    method_stats = {
        method: {
            "mean_error_pct": round(np.mean(errs), 2),
            "std_dev": round(np.std(errs, ddof=1), 2) if len(errs) > 1 else 0.0
        }
        for method, errs in global_method_errors.items()
    }

    return raw_data, error_by_test_case, error_summary, method_stats

    # detailed_rows = []
    # function_summary_rows = []
    # method_overall_rows = []

    # # Para calcular los errores
    # accumulated_errors = defaultdict(lambda: defaultdict(list))  # {fname: {method: [errors]}}
    # global_errors = defaultdict(list)  # {method: [all errors across functions]}

    # for fname, intervals_dict in new_results.items():
    #     for interval_str, impls in intervals_dict.items():
    #         # Calcular el rango (Max - Min) por implementación
    #         impls.pop('Ibex')
    #         ranges = {
    #             impl: data["Max"] - data["Min"]
    #             for impl, data in impls.items()
    #         }

    #         best_range = min(ranges.values()) or 1e-12  # evitar división por cero

    #         # Errores relativos (%)
    #         error_percentages = {
    #             impl: 100 * abs(r - best_range) / best_range
    #             for impl, r in ranges.items()
    #         }

    #         print('error percentajes', error_percentages)

    #         # Acumular errores por función y globales
    #         for method, err in error_percentages.items():
    #             accumulated_errors[fname][method].append(err)
    #             global_errors[method].append(err)

    #         # Construir fila detallada por intervalo
    #         row = {
    #             "fname": fname,
    #             "interval": interval_str,
    #         }

    #         values = []
    #         for method in sorted(error_percentages.keys()):
    #             val = error_percentages[method]
    #             row[f"% error {method}"] = val
    #             values.append(val)

    #         row["desv estandar"] = np.std(values)
    #         detailed_rows.append(row)

    # # Construir resumen por función
    # for fname, method_errors in accumulated_errors.items():
    #     row = {"fname": fname, "interval": "PROMEDIO"}
    #     method_means = []

    #     for method in sorted(method_errors.keys()):
    #         mean = sum(method_errors[method]) / len(method_errors[method])
    #         row[f"% error {method}"] = mean
    #         method_means.append(mean)

    #     row["desv estandar"] = np.std(method_means)
    #     function_summary_rows.append(row)

    # # Construir resumen global por método
    # for method, all_errors in global_errors.items():
    #     row = {
    #         "metodo": method,
    #         "promedio global de error": sum(all_errors) / len(all_errors),
    #         "desviación estándar entre funciones": np.std(all_errors)
    #     }
    #     method_overall_rows.append(row)

    # # ---------- Cuarto DataFrame: Datos crudos ----------
    # raw_rows = []

    # for fname, intervals_dict in new_results.items():
    #     for interval_str, impls in intervals_dict.items():
    #         for impl_name, data in impls.items():
    #             raw_rows.append({
    #                 "fname": fname,
    #                 "f": data["Function"],
    #                 "interval": data["Interval"],
    #                 "implementation": data["Implementation"],
    #                 "min": data["Min"],
    #                 "max": data["Max"]
    #             })

    # return detailed_rows, function_summary_rows, method_overall_rows, raw_rows

def export_to_excel(data, filename="output.xlsx"):
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)

    # Cargar archivo con openpyxl para modificar formato
    wb = load_workbook(filename)
    ws = wb.active

    # Congelar la primera fila
    ws.freeze_panes = "A2"

    # Aplicar estilo y ajustar ancho de columnas
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Establecer ancho de columna con un poco de margen extra
        ws.column_dimensions[col_letter].width = max_length + 2

        # Estilizar encabezado (negrita)
        header_cell = ws[f"{col_letter}1"]
        header_cell.font = Font(bold=True)

    wb.save(filename)

def str_to_func(func_str, n_vars):
    """
    Convierte la función en un lambda
    """
    vars = sp.symbols(f'x:{n_vars}')
    print(vars, func_str)

    for i in range(n_vars):
        func_str = func_str.replace(f'x[{i}]', f'x{i}')
    
    # Reemplazar ^ por ** para exponenciación
    func_str = func_str.replace('^', '**')

    expr = sp.sympify(func_str)
    return sp.lambdify(vars, expr, 'numpy')

def init_boxes(global_x, intervals):
    Affine.x_ref = global_x
    Affine.envs = dict()

    n = global_x.size()

    linear_vars = []
    for i in range(n):
        a = [1 if j == i else 0 for j in range(n)]
        linear_vars.append(LinearExpression(a,0, intervals))

    envs = []
    for i, var in enumerate(linear_vars):
        Affine.envs[i] = LinearEnvelope(is_var=True, linear_expr=var)
        envs.append(Affine.envs[i])

    return linear_vars, envs


def generate_interval(minval=1, maxval=100):
    """
    Genera un intervalo aleatorio [min, max] con un valor semilla dado.
    """
    random.seed()
    a = random.randint(minval, maxval)
    b = random.randint(minval, maxval)

    if a == b:
        b += 1
        
    return Interval(min(a, b), max(a, b)), IntervalTraditional(min(a, b), max(a, b))